"""
Safety Stock Recommendation Tool
Plant-level safety stock analysis from SAP MARC + MB51 extracts.

Run with:  streamlit run app.py
"""

import io
import os
from datetime import datetime

import numpy as np
import pandas as pd
import streamlit as st

# ---------------------------------------------------------------------------
# Configuration: column mapping
# If a future SAP extract uses different headers, adjust here only.
# ---------------------------------------------------------------------------
MARC_COLS = {
    "material": "Material",
    "plant": "Plnt",
    "mrp_type": "Typ",
    "mrp_controller": "MRPCn",
    "safety_stock": "Safety Stock",
    "fix_lot": "Fix. lot size",
    "pdt": "PDT",
    "grt": "GRT",
    "trlt": "TRLT",
    "abc": "ABC",
}

MB51_COLS = {
    "material": "Material",
    "description": "Material Description",
    "plant": "Plant",
    "movement_type": "Movement Type",
    "posting_date": "Posting Date",
    "qty": "Qty in unit of entry",
    "unit": "Unit of Entry",
}

Z_SCORES = {
    "90% (Z = 1.28)": 1.28,
    "95% (Z = 1.65)": 1.65,
    "97.5% (Z = 1.96)": 1.96,
    "99% (Z = 2.33)": 2.33,
}

# ---------------------------------------------------------------------------
# Page setup
# ---------------------------------------------------------------------------
# Messer brand palette
MESSER_BLUE = "#134395"
MESSER_RED = "#E3312A"

st.set_page_config(
    page_title="Safety Stock Recommendation Tool",
    page_icon="📦",
    layout="wide",
)

# Brand styling (no external font import, so it works on offline/locked-down networks).
# We prefer Open Sans if it is installed locally, otherwise fall back to system sans-serif.
# The font rule targets the app container but explicitly excludes icon spans, so it cannot
# break Material icon ligatures (e.g. the file-uploader button).
st.markdown(
    f"""
    <style>
    [data-testid="stAppViewContainer"], [data-testid="stSidebar"] {{
        font-family: 'Open Sans', 'Segoe UI', Helvetica, Arial, sans-serif;
    }}
    [data-testid="stAppViewContainer"] [class*="material-icons"],
    [data-testid="stSidebar"] [class*="material-icons"] {{
        font-family: 'Material Symbols Rounded', 'Material Icons' !important;
    }}
    h1, h2, h3 {{ color: {MESSER_BLUE}; }}
    [data-testid="stMetricValue"] {{ color: {MESSER_BLUE}; }}
    </style>
    """,
    unsafe_allow_html=True,
)

_logo_col, _title_col = st.columns([1, 5])
with _logo_col:
    if os.path.exists("messer_logo.png"):
        try:
            st.image("messer_logo.png", width=150)
        except Exception:
            # Corrupt or unreadable image: skip the logo rather than crash the app
            pass
with _title_col:
    st.title("Safety Stock Recommendation Tool")
st.caption(
    "Upload SAP MARC and MB51 extracts, tune the parameters, and generate "
    "safety stock recommendations per material."
)

# ---------------------------------------------------------------------------
# Sidebar: parameters
# ---------------------------------------------------------------------------
with st.sidebar:
    st.header("⚙️ Parameters")

    service_label = st.selectbox(
        "Service level",
        list(Z_SCORES.keys()),
        index=1,
        help="Target probability of not stocking out during the replenishment cycle.",
    )
    z_value = Z_SCORES[service_label]

    mvt_choice = st.radio(
        "Demand signal (movement types)",
        ["641 + 601 combined", "641 only (STO)", "601 only (Sales GI)"],
        help="Which outbound movements count as demand.",
    )
    if mvt_choice == "641 only (STO)":
        mvt_types = [641]
    elif mvt_choice == "601 only (Sales GI)":
        mvt_types = [601]
    else:
        mvt_types = [641, 601]

    lt_source = st.radio(
        "Lead time source",
        ["TRLT (Total Replenishment Lead Time)", "PDT + GRT"],
        help=(
            "TRLT covers the full replenishment cycle. PDT+GRT covers "
            "delivery plus goods receipt processing only."
        ),
    )

    dos_days = st.slider(
        "Days of Supply benchmark (days)",
        min_value=5,
        max_value=60,
        value=15,
        step=5,
        help="Simple buffer expressed as days of average demand. "
        "The recommendation takes the higher of statistical SS and this DoS value.",
    )

    delta_threshold = st.slider(
        "Delta alignment threshold (CYL)",
        min_value=1,
        max_value=20,
        value=5,
        help="Materials within ± this many cylinders of current SS are considered aligned.",
    )

    st.divider()
    st.caption(
        "Methodology: Statistical SS = Z × σ(monthly demand) × √(lead time in months). "
        "Recommendation = max(Statistical SS, DoS benchmark), rounded up."
    )

# ---------------------------------------------------------------------------
# File upload
# ---------------------------------------------------------------------------
col1, col2 = st.columns(2)
with col1:
    marc_file = st.file_uploader("MARC extract (.xlsx)", type=["xlsx"], key="marc")
with col2:
    mb51_file = st.file_uploader("MB51 extract (.xlsx)", type=["xlsx"], key="mb51")


def validate_columns(df: pd.DataFrame, required: dict, name: str) -> list:
    """Return list of missing columns."""
    return [v for v in required.values() if v not in df.columns]


@st.cache_data(show_spinner=False)
def load_excel(file_bytes: bytes) -> pd.DataFrame:
    return pd.read_excel(io.BytesIO(file_bytes))


# ---------------------------------------------------------------------------
# Core analysis
# ---------------------------------------------------------------------------
def run_analysis(
    marc: pd.DataFrame,
    mb51: pd.DataFrame,
    z: float,
    movement_types: list,
    lead_time_source: str,
    dos_benchmark_days: int,
) -> tuple[pd.DataFrame, dict]:
    marc = marc.copy()
    mb51 = mb51.copy()

    marc[MARC_COLS["material"]] = marc[MARC_COLS["material"]].astype(str)
    mb51[MB51_COLS["material"]] = mb51[MB51_COLS["material"]].astype(str)
    mb51["Qty_abs"] = mb51[MB51_COLS["qty"]].abs()
    mb51[MB51_COLS["posting_date"]] = pd.to_datetime(mb51[MB51_COLS["posting_date"]])

    demand = mb51[mb51[MB51_COLS["movement_type"]].isin(movement_types)].copy()
    if demand.empty:
        return pd.DataFrame(), {}

    demand["YearMonth"] = demand[MB51_COLS["posting_date"]].dt.to_period("M")
    monthly = (
        demand.groupby([MB51_COLS["material"], "YearMonth"])["Qty_abs"]
        .sum()
        .reset_index()
    )

    date_min = demand[MB51_COLS["posting_date"]].min()
    date_max = demand[MB51_COLS["posting_date"]].max()
    all_periods = pd.period_range(date_min.to_period("M"), date_max.to_period("M"), freq="M")

    results = []
    for _, mrow in marc.iterrows():
        mat = mrow[MARC_COLS["material"]]
        trlt = mrow.get(MARC_COLS["trlt"], 0) or 0
        pdt = mrow.get(MARC_COLS["pdt"], 0) or 0
        grt = mrow.get(MARC_COLS["grt"], 0) or 0
        current_ss = mrow.get(MARC_COLS["safety_stock"], 0) or 0
        mrp_type = mrow.get(MARC_COLS["mrp_type"], "")
        mrp_cn = mrow.get(MARC_COLS["mrp_controller"], "")
        abc = mrow.get(MARC_COLS["abc"], "")

        desc_series = mb51.loc[
            mb51[MB51_COLS["material"]] == mat, MB51_COLS["description"]
        ].dropna()
        desc = desc_series.iloc[0] if not desc_series.empty else ""

        if lead_time_source.startswith("TRLT"):
            lt_days = trlt if trlt > 0 else (pdt + grt)
            lt_note = "TRLT=0; used PDT+GRT as fallback" if trlt == 0 and (pdt + grt) > 0 else ""
        else:
            lt_days = pdt + grt
            lt_note = ""

        mat_monthly = monthly[monthly[MB51_COLS["material"]] == mat]

        if mat_monthly.empty:
            results.append(
                {
                    "Material": mat,
                    "Description": desc,
                    "ABC": abc,
                    "MRP Type": mrp_type,
                    "MRP Controller": mrp_cn,
                    "Lead Time (days)": int(lt_days),
                    "Avg Monthly Demand": 0.0,
                    "Std Dev Monthly": 0.0,
                    "CV": np.nan,
                    "Stat SS": np.nan,
                    f"DoS {dos_benchmark_days}-day SS": np.nan,
                    "Recommended SS": 0,
                    "Current SS (MARC)": int(current_ss),
                    "Delta": 0 - int(current_ss),
                    "Note": "No demand history in selected movement types",
                }
            )
            continue

        series = (
            mat_monthly.set_index("YearMonth")["Qty_abs"]
            .reindex(all_periods, fill_value=0)
        )
        avg_m = series.mean()
        std_m = series.std(ddof=1)
        cv = std_m / avg_m if avg_m > 0 else 0

        lt_months = lt_days / 30
        stat_ss = z * std_m * np.sqrt(lt_months) if lt_months > 0 else z * std_m
        dos_ss = (avg_m / 30) * dos_benchmark_days
        recommended = int(np.ceil(max(stat_ss, dos_ss)))

        results.append(
            {
                "Material": mat,
                "Description": desc,
                "ABC": abc,
                "MRP Type": mrp_type,
                "MRP Controller": mrp_cn,
                "Lead Time (days)": int(lt_days),
                "Avg Monthly Demand": round(avg_m, 1),
                "Std Dev Monthly": round(std_m, 1),
                "CV": round(cv, 2),
                "Stat SS": round(stat_ss, 1),
                f"DoS {dos_benchmark_days}-day SS": round(dos_ss, 1),
                "Recommended SS": recommended,
                "Current SS (MARC)": int(current_ss),
                "Delta": recommended - int(current_ss),
                "Note": lt_note,
            }
        )

    df = pd.DataFrame(results).sort_values(
        ["ABC", "Avg Monthly Demand"], ascending=[True, False]
    )

    meta = {
        "date_min": date_min,
        "date_max": date_max,
        "n_months": len(all_periods),
        "n_materials": len(df),
        "n_no_history": int((df["Note"].str.contains("No demand", na=False)).sum()),
    }
    return df, meta


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------
def build_excel(df: pd.DataFrame, meta: dict, params: dict, delta_thr: int) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "SS Recommendations"

    # Messer brand colors
    messer_blue = "134395"
    messer_blue_tint = "5A7BB5"  # 30% white tint of Messer Blue, for header row
    messer_red = "E3312A"
    brand_font = "Open Sans"

    header_fill = PatternFill("solid", fgColor=messer_blue)
    sub_fill = PatternFill("solid", fgColor=messer_blue_tint)
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    orange_fill = PatternFill("solid", fgColor="FCE4D6")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    grey_fill = PatternFill("solid", fgColor="EEF2F8")  # light blue-grey stripe

    white_font = Font(color="FFFFFF", bold=True, name=brand_font, size=10)
    bold_font = Font(bold=True, name=brand_font, size=10)
    reg_font = Font(name=brand_font, size=10)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    n_cols = len(df.columns)
    last_col = get_column_letter(n_cols)

    # Title row: white background, Messer Blue text, Messer Red rule underneath
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = (
        f"Safety Stock Recommendations | {params['service']} | "
        f"Demand: {params['mvt']} | Lead time: {params['lt']} | "
        f"Window: {meta['date_min']:%b %Y} - {meta['date_max']:%b %Y}"
    )
    ws["A1"].font = Font(color=messer_blue, bold=True, name=brand_font, size=12)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    red_rule = Side(style="thick", color=messer_red)
    for c in range(1, n_cols + 1):
        ws.cell(row=1, column=c).border = Border(bottom=red_rule)
    ws.row_dimensions[1].height = 40

    # Embed logo in top-left if available
    try:
        from openpyxl.drawing.image import Image as XLImage

        if os.path.exists("messer_logo.png"):
            img = XLImage("messer_logo.png")
            img.height = 42
            img.width = int(42 * 1690 / 745)
            ws.add_image(img, "A1")
    except Exception:
        pass

    for c_idx, h in enumerate(df.columns, 1):
        cell = ws.cell(row=2, column=c_idx, value=h)
        cell.font = white_font
        cell.fill = sub_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[2].height = 30

    delta_col = list(df.columns).index("Delta") + 1
    cv_col = list(df.columns).index("CV") + 1
    rec_col = list(df.columns).index("Recommended SS") + 1

    for r_idx, (_, row) in enumerate(df.iterrows(), 3):
        no_hist = "No demand" in str(row["Note"])
        base = orange_fill if no_hist else (white_fill if r_idx % 2 == 1 else grey_fill)
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = None if (isinstance(val, float) and np.isnan(val)) else val
            cell.font = reg_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = base
        if not no_hist:
            d = row["Delta"]
            dc = ws.cell(row=r_idx, column=delta_col)
            if d > delta_thr:
                dc.fill = red_fill
            elif d < -delta_thr:
                dc.fill = green_fill
            else:
                dc.fill = yellow_fill
            cv = row["CV"]
            if isinstance(cv, (int, float)) and not np.isnan(cv):
                cc = ws.cell(row=r_idx, column=cv_col)
                if cv >= 2.0:
                    cc.fill = red_fill
                elif cv >= 1.0:
                    cc.fill = yellow_fill
        ws.cell(row=r_idx, column=rec_col).font = bold_font

    widths = [12, 38, 6, 10, 12, 14, 20, 16, 10, 14, 16, 16, 18, 12, 34]
    for i in range(1, n_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1] if i <= len(widths) else 16
    ws.freeze_panes = "C3"

    # Methodology sheet
    ls = wb.create_sheet("Notes & Methodology")
    notes = [
        ("Safety Stock Analysis | Methodology Notes", True),
        ("", False),
        (f"Generated: {datetime.now():%Y-%m-%d %H:%M}", False),
        (f"Service level: {params['service']}", False),
        (f"Demand signal: {params['mvt']}", False),
        (f"Lead time source: {params['lt']}", False),
        (f"DoS benchmark: {params['dos']} days", False),
        (f"Observation window: {meta['date_min']:%b %Y} to {meta['date_max']:%b %Y} ({meta['n_months']} months)", False),
        ("", False),
        ("FORMULAS", True),
        ("Statistical SS = Z × σ(monthly demand) × √(lead time in months)", False),
        ("DoS SS = (avg monthly demand / 30) × benchmark days", False),
        ("Recommended SS = ceiling(max(Statistical SS, DoS SS))", False),
        ("Zero-demand months are included in the standard deviation.", False),
        ("", False),
        ("COLOR CODING", True),
        (f"Delta red: recommended > current by more than {delta_thr} (understocked)", False),
        (f"Delta green: recommended < current by more than {delta_thr} (overstocked)", False),
        (f"Delta yellow: within ±{delta_thr} (aligned)", False),
        ("CV red ≥ 2.0: very high variability | CV yellow ≥ 1.0: moderate", False),
        ("Orange rows: no demand history in the selected movement types.", False),
    ]
    ls.column_dimensions["A"].width = 90
    for i, (text, is_bold) in enumerate(notes, 1):
        cell = ls.cell(row=i, column=1, value=text)
        cell.font = Font(bold=is_bold, name=brand_font, size=10, color=messer_blue if is_bold else "000000")
        if is_bold and text:
            cell.fill = PatternFill("solid", fgColor="E3EAF6")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Main flow
# ---------------------------------------------------------------------------
if marc_file and mb51_file:
    marc_df = load_excel(marc_file.getvalue())
    mb51_df = load_excel(mb51_file.getvalue())

    missing_marc = validate_columns(marc_df, MARC_COLS, "MARC")
    missing_mb51 = validate_columns(mb51_df, MB51_COLS, "MB51")

    if missing_marc or missing_mb51:
        if missing_marc:
            st.error(f"MARC file is missing required columns: {missing_marc}")
        if missing_mb51:
            st.error(f"MB51 file is missing required columns: {missing_mb51}")
        st.info(
            "If your SAP extract uses different column headers, update the "
            "MARC_COLS / MB51_COLS mapping at the top of app.py."
        )
        st.stop()

    df, meta = run_analysis(
        marc_df, mb51_df, z_value, mvt_types,
        lt_source, dos_days,
    )

    if df.empty:
        st.warning("No demand records found for the selected movement types.")
        st.stop()

    # --- KPI cards ---
    active = df[~df["Note"].str.contains("No demand", na=False)]
    n_under = int((active["Delta"] > delta_threshold).sum())
    n_over = int((active["Delta"] < -delta_threshold).sum())
    n_aligned = len(active) - n_under - n_over

    st.subheader("Summary")
    k1, k2, k3, k4, k5 = st.columns(5)
    k1.metric("Materials analyzed", meta["n_materials"])
    k2.metric("Understocked", n_under, help=f"Delta > +{delta_threshold} CYL")
    k3.metric("Aligned", n_aligned, help=f"Within ±{delta_threshold} CYL")
    k4.metric("Overstocked", n_over, help=f"Delta < -{delta_threshold} CYL")
    k5.metric("No demand history", meta["n_no_history"])

    st.caption(
        f"Demand window: {meta['date_min']:%b %Y} to {meta['date_max']:%b %Y} "
        f"({meta['n_months']} months) | Zero-demand months included in variability."
    )

    # --- Results table with styling ---
    st.subheader("Recommendations")

    seg_filter = st.multiselect(
        "Filter by segment",
        ["Understocked", "Aligned", "Overstocked", "No demand history"],
        default=[],
        placeholder="Show all",
    )

    def segment(row):
        if "No demand" in str(row["Note"]):
            return "No demand history"
        if row["Delta"] > delta_threshold:
            return "Understocked"
        if row["Delta"] < -delta_threshold:
            return "Overstocked"
        return "Aligned"

    df_view = df.copy()
    df_view["Segment"] = df_view.apply(segment, axis=1)
    if seg_filter:
        df_view = df_view[df_view["Segment"].isin(seg_filter)]

    def style_delta(val):
        if pd.isna(val):
            return ""
        if val > delta_threshold:
            return "background-color: #FFC7CE"
        if val < -delta_threshold:
            return "background-color: #C6EFCE"
        return "background-color: #FFEB9C"

    def style_cv(val):
        if pd.isna(val):
            return ""
        if val >= 2.0:
            return "background-color: #FFC7CE"
        if val >= 1.0:
            return "background-color: #FFEB9C"
        return ""

    styled = (
        df_view.drop(columns=["Segment"])
        .style
        .applymap(style_delta, subset=["Delta"])
        .applymap(style_cv, subset=["CV"])
        .format(precision=1, na_rep="N/A")
    )
    st.dataframe(styled, use_container_width=True, height=520)

    # --- Per-material drill down ---
    st.subheader("Material drill-down")
    mat_options = df[~df["Note"].str.contains("No demand", na=False)]["Material"].tolist()
    sel_mat = st.selectbox("Select a material to view its monthly demand pattern", mat_options)

    if sel_mat:
        mb51_df["Material"] = mb51_df[MB51_COLS["material"]].astype(str)
        d = mb51_df[
            (mb51_df["Material"] == sel_mat)
            & (mb51_df[MB51_COLS["movement_type"]].isin(mvt_types))
        ].copy()
        d["Qty_abs"] = d[MB51_COLS["qty"]].abs()
        d[MB51_COLS["posting_date"]] = pd.to_datetime(d[MB51_COLS["posting_date"]])
        d["YearMonth"] = d[MB51_COLS["posting_date"]].dt.to_period("M").dt.to_timestamp()
        chart_data = d.groupby("YearMonth")["Qty_abs"].sum()

        full_range = pd.date_range(
            meta["date_min"].to_period("M").to_timestamp(),
            meta["date_max"].to_period("M").to_timestamp(),
            freq="MS",
        )
        chart_data = chart_data.reindex(full_range, fill_value=0)

        row = df[df["Material"] == sel_mat].iloc[0]
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Avg monthly demand", f"{row['Avg Monthly Demand']:.1f}")
        c2.metric("CV", f"{row['CV']:.2f}")
        c3.metric("Recommended SS", int(row["Recommended SS"]))
        c4.metric("Current SS", int(row["Current SS (MARC)"]))

        st.bar_chart(chart_data, height=280)

    # --- Export ---
    st.subheader("Export")
    params = {
        "service": service_label,
        "mvt": mvt_choice,
        "lt": lt_source,
        "dos": dos_days,
    }
    excel_bytes = build_excel(df, meta, params, delta_threshold)
    st.download_button(
        "⬇️ Download formatted Excel",
        data=excel_bytes,
        file_name=f"SS_Recommendations_{datetime.now():%Y%m%d}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

else:
    st.info("Upload both files above to run the analysis.")
    with st.expander("Expected file formats"):
        st.markdown(
            """
**MARC extract** must contain at minimum:
`Material`, `Plnt`, `Typ`, `MRPCn`, `Safety Stock`, `PDT`, `GRT`, `TRLT`, `ABC`, `Fix. lot size`

**MB51 extract** must contain at minimum:
`Material`, `Material Description`, `Plant`, `Movement Type`, `Posting Date`, `Qty in unit of entry`, `Unit of Entry`

If your extracts use different column headers, edit the `MARC_COLS` and
`MB51_COLS` dictionaries at the top of `app.py`.
"""
        )
